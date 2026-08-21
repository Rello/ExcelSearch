from datetime import date
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook as load_openpyxl_workbook

from excel_search.workbook import export_dataframe, load_workbook


def test_load_workbook_lists_and_selects_sheets(tmp_path: Path) -> None:
    workbook = tmp_path / "example.xlsx"
    with pd.ExcelWriter(workbook, engine="openpyxl") as writer:
        pd.DataFrame({"Name": ["Anna"]}).to_excel(writer, sheet_name="Kunden", index=False)
        pd.DataFrame({"Artikel": ["Tisch"]}).to_excel(writer, sheet_name="Produkte", index=False)

    sheets, selected, dataframe = load_workbook(workbook, "Produkte")

    assert sheets == ["Kunden", "Produkte"]
    assert selected == "Produkte"
    assert dataframe.to_dict("records") == [{"Artikel": "Tisch"}]


def test_unsupported_file_type_is_rejected(tmp_path: Path) -> None:
    with pytest.raises(ValueError, match="Unterstützt werden"):
        load_workbook(tmp_path / "example.csv")


def test_export_xlsx_round_trip(tmp_path: Path) -> None:
    destination = tmp_path / "result.xlsx"
    dataframe = pd.DataFrame({"Name": ["Müller"]})

    export_dataframe(dataframe, destination)

    assert pd.read_excel(destination, dtype=str).to_dict("records") == [{"Name": "Müller"}]


def test_export_csv_uses_bom_and_selected_separator(tmp_path: Path) -> None:
    destination = tmp_path / "result.csv"

    export_dataframe(
        pd.DataFrame({"Name": ["Müller"], "Ort": ["Köln"]}),
        destination,
        csv_separator=";",
    )

    raw = destination.read_bytes()
    assert raw.startswith(b"\xef\xbb\xbf")
    assert "Name;Ort" in raw.decode("utf-8-sig")


def test_exports_format_dates_as_day_month_year(tmp_path: Path) -> None:
    dataframe = pd.DataFrame({"Datum": [date(2026, 8, 21)]})
    csv_destination = tmp_path / "dates.csv"
    xlsx_destination = tmp_path / "dates.xlsx"

    export_dataframe(dataframe, csv_destination)
    export_dataframe(dataframe, xlsx_destination)

    assert "21.08.2026" in csv_destination.read_text(encoding="utf-8-sig")
    workbook = load_openpyxl_workbook(xlsx_destination)
    assert workbook.active["A2"].number_format == "DD.MM.YYYY"


def test_xls_reader_dependency_is_available() -> None:
    import xlrd

    assert xlrd.__version__ == "2.0.2"
